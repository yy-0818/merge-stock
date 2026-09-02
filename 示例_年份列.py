#!/usr/bin/env python3
"""
年份列使用示例
演示如何在实际项目中使用带年份列的库存表
"""

import openpyxl
from pathlib import Path

def create_sample_with_year():
    """创建带年份列的示例库存表"""
    
    print("=" * 60)
    print("创建示例：带年份列的库存表")
    print("=" * 60)
    
    # 创建示例文件
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 表头 - 年份列在第二列（客户名之后）
    ws.append([
        '客户名', '年份', '型号', '类型', 
        '1级', '2级', 
        'D1', 'D2', 'D3', 'D4', 'D5',
        'A', 'A1', 'A2',
        '辅助列'
    ])
    
    # 数据行 - 多个年份的数据
    data = [
        ['东方陶瓷', 2024, 'XM-001', '瓷砖', 1500, 800, 100, 150, 200, 0, 50, 300, 100, 50, '显示'],
        ['东方陶瓷', 2025, 'XM-001', '瓷砖', 1200, 600, 80, 120, 180, 0, 40, 250, 80, 40, '显示'],
        ['东方陶瓷', 2024, 'XM-002', '墙砖', 2000, 1000, 150, 200, 250, 100, 80, 400, 150, 80, '显示'],
        ['东方陶瓷', 2025, 'XM-002', '墙砖', 1800, 900, 130, 180, 230, 90, 70, 350, 130, 70, '显示'],
        ['西湖建材', 2024, 'XM-003', '地砖', 3000, 1500, 200, 300, 400, 150, 100, 500, 200, 100, '显示'],
        ['西湖建材', 2025, 'XM-003', '地砖', 2800, 1400, 180, 280, 380, 140, 90, 480, 180, 90, '显示'],
        ['', '', '合计', '合计', 12300, 6200, 840, 1230, 1640, 480, 430, 2280, 840, 430, ''],
    ]
    
    for row in data:
        ws.append(row)
    
    # 保存示例文件
    sample_path = Path('/Users/kaiyuan/work/stock/示例_带年份列.xlsx')
    wb.save(sample_path)
    print(f"\n✓ 示例文件已创建: {sample_path}")
    
    # 打印表格预览
    print("\n表格预览（前5列）：")
    print("-" * 60)
    wb_read = openpyxl.load_workbook(sample_path)
    ws_read = wb_read.active
    
    for i, row in enumerate(ws_read.iter_rows(values_only=True), 1):
        if i == 1:
            print(f"{'  '.join(str(v)[:10].ljust(10) for v in row[:5])}")
            print("-" * 60)
        else:
            print(f"{'  '.join(str(v)[:10].ljust(10) if v else ' '*10 for v in row[:5])}")
        
        if i >= 8:
            break
    
    wb_read.close()
    
    # 说明
    print("\n" + "=" * 60)
    print("使用说明：")
    print("=" * 60)
    print("""
1. 表头必须包含"年份"列（名称区分大小写不敏感）
2. 年份值应为数字（如 2024, 2025）
3. 年份列可以放在任意位置，脚本会自动识别
4. 合并时年份列会应用数据列样式（右对齐、底色、数字格式）
5. 年份不参与排序，仅作为维度展示

常见布局方案：
  方案A：客户名 → 年份 → 型号 → 类型 → 数据列...
  方案B：客户名 → 型号 → 类型 → 年份 → 数据列...
  方案C：客户名 → 型号 → 年份 → 类型 → 数据列...

所有方案都能正确处理！
    """)
    
    print("=" * 60)
    print("运行合并：")
    print("=" * 60)
    print("""
方式1 - GUI：
    python3 gui.py
    然后在界面中选择包含年份列的源表

方式2 - CLI：
    python3 merge_stock_files.py \\
        --src <源目录> \\
        --index <分类表> \\
        --out <输出目录>

方式3 - 测试：
    python3 test_year_column.py
    """)
    
    wb.close()

if __name__ == "__main__":
    create_sample_with_year()
