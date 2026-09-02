"""
按 分类.xlsx 把 stock/2026-7-17/ 下的子表拼接成 A 列命名的大表，并按组归档。

输出结构（每运行一次创建一个新的时间戳目录）：
  <OUTPUT_DIR>/<YYYY-MM-DD_HH-MM-SS>/
    <A列名>/
      <A列名>.xlsx       ← 该组所有子表纵向拼接（保留各自表头，按辅助列过滤）
      <子表名>.xlsx      ← 分类表 C 列起的源子表
      ...
    ...

每次运行自动创建新的时间戳目录，保留历史运行结果。

规则：
- 每个子表的最后两行通常为 (空行, 合计行)；合计行的特征是「型号 == '合计'」
  兜底：当某张子表没有此字样时，取末尾倒数第二个非空数据行
- 辅助列（最后一列）= 显示 → 保留；不显示 → 丢弃；合计行保留
- 写表时统一丢掉辅助列
- 行内容：保留源表真实宽度，不做补齐/截断；完全空行跳过

样式（在合并文件上自动应用）：
- 表头行：深蓝底 #4472C4 + 白字 + 粗体 + 居中
- 数据行：G 列起按列固定底色（36 色黄金角分布，每列一眼区分其归属档位 1级/2级/D1..D22/A/A1..A12）
- 空数据列不涂色；同一列内空单元格也不涂色，让其它已涂色的格子视觉一致
- 合计行：浅黄 #FFE699 + 粗体（覆盖整行）
- 全表灰色细边框
- 冻结第 1 行
- 列宽按内容自动估算，上限 50

拆分逻辑（自动）：
- 列数不再手动指定：扫描组内所有源子表的「真实最大列数」作为合并目标列数
- 空列自动移除：若某一列在整组合并数据中**所有行均为空**（包含表头），则直接剔除该列，
  不写入输出表，达到压缩列数优化展示的效果
- 仍保留原有行为：辅助列（显示/不显示控制）、换算率/备注列始终被剔除

本模块提供：
  load_config()         -> dict
  process(cfg, progress_cb) -> ProcessResult   (供 GUI / CLI 共用)
  main()                                          (CLI 入口)
"""
from __future__ import annotations

import json
import os
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ------------------- 路径解析 -------------------
def exe_dir() -> Path:
    """脚本运行/被打包后所在目录：开发时是脚本所在目录，打包后是 exe 同级目录。"""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent.resolve()


# 路径全部留空：打包后的二进制必须依赖同级目录的 config.json 才能运行
DEFAULT_CONFIG = {
    "src_dir": "",
    "index_file": "",
    "output_dir": "",
}


def load_config() -> dict:
    """优先读 <exe_dir>/config.json，未提供字段用默认值。"""
    cfg_path = exe_dir() / "config.json"
    cfg = dict(DEFAULT_CONFIG)
    if cfg_path.exists():
        try:
            user_cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
            if isinstance(user_cfg, dict):
                cfg.update({k: v for k, v in user_cfg.items() if k in DEFAULT_CONFIG})
        except Exception as e:
            print(f"[警告] config.json 解析失败，已忽略: {e}")
    else:
        example = exe_dir() / "config.example.json"
        if not example.exists():
            example.write_text(
                json.dumps(DEFAULT_CONFIG, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
    return cfg


def save_config(cfg: dict, path: Path | None = None) -> Path:
    """把 cfg 写到 <exe_dir>/config.json（GUI 持久化用）。"""
    p = path or (exe_dir() / "config.json")
    p.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
    return p


# ------------------- 数据结构 -------------------
class UnsafeOutputDirError(Exception):
    """output_dir 指向危险路径(如 ~/Desktop)时抛出,需要 cfg['allow_unsafe_output']=True 才放行。"""


def _resolve_real(p: Path) -> Path:
    """解析 symlink、相对路径,得到真实绝对路径。用于和家目录/根目录比较。"""
    return p.expanduser().resolve()


_DANGEROUS_OUTPUT_DIRS: tuple[tuple[Path, str], ...] = (
    (Path.home(), "用户主目录 (~)"),
    (Path.home() / "Desktop", "桌面 (~)"),
    (Path.home() / "Documents", "文档 (~)"),
    (Path.home() / "Downloads", "下载 (~)"),
    (Path("/"), "根目录 (/)"),
)


def _is_dangerous_output_dir(p: Path) -> str | None:
    """返回命中的危险路径说明;否则 None。仅在严格相等(resolve 之后)命中。"""
    real = _resolve_real(p)
    for dangerous, label in _DANGEROUS_OUTPUT_DIRS:
        try:
            if real == dangerous.resolve():
                return label
        except Exception:
            continue
    return None


def _prepare_output_dir(output_dir: Path, allow_unsafe: bool = False) -> Path:
    """为每次运行创建新的带时间戳的输出目录。

    策略:
    1. resolve 后,若等于 ~/Desktop ~/Documents ~/Downloads ~ / 之一 → 抛 UnsafeOutputDirError
       (除非显式传入 allow_unsafe=True)
    2. 在 output_dir 下创建新的带时间戳目录: YYYY-MM-DD_HH-MM-SS
    3. 返回实际使用的目录路径
    """
    if not allow_unsafe:
        label = _is_dangerous_output_dir(output_dir)
        if label is not None:
            raise UnsafeOutputDirError(
                f"output_dir 指向危险位置: {label}\n"
                f"  → 为了避免误删整个 {label},脚本拒绝使用此路径作为输出。\n"
                f"  请新建一个子目录(例如 ~/work/stock/_output)再设置 output_dir。"
            )

    output_dir.mkdir(parents=True, exist_ok=True)

    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    run_dir = output_dir / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


@dataclass
class ProcessResult:
    groups_total: int = 0
    groups_merged: int = 0
    rows_total: int = 0                # 过滤后写出的行数（不含"不显示"被丢的行）
    rows_filtered_total: int = 0       # 过滤掉多少行（"不显示"）
    files_copied: int = 0
    missing_files: list[str] = field(default_factory=list)
    log: list[str] = field(default_factory=list)
    actual_output_dir: Path | None = None  # 实际使用的输出目录（含时间戳）

    @property
    def ok(self) -> bool:
        return self.groups_merged > 0


# ------------------- 样式常量 (professional client-facing preset) -------------------
# 设计目标: 报表可直接邮件给客户 → 干净、专业、易读、有视觉层次
#
# 视觉系统:
#   - 表头:深海军蓝底 + 白色粗体, 高度 32px, 冻结首行
#   - 数据行:相邻 section 用浅/深交替(斑马纹), 隔 5 列轻微色相切换增强可读性
#   - 数字右对齐(便于上下比较);文本列(型号/类型)居中
#   - 合计行:温暖橙黄底 + 粗体 + 上方双线,视觉跳出
#   - 全表细线边框, 颜色柔和不刺眼
#   - 客户名首列(A):浅灰底,视觉锚定 "这是谁的库存"
#   - 行高 22px, 列宽自适应 + padding 2
#
HEAD_FILL = PatternFill("solid", fgColor="1F3864")                # 深海军蓝
HEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEAD_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")       # 浅黄(温暖醒目,适合数字列)
TOTAL_FONT = Font(name="Calibri", bold=True, color="7F4E00", size=11)
TOTAL_ALIGN = Alignment(horizontal="center", vertical="center")

# 数据行
DATA_FONT = Font(name="Calibri", size=11, color="1F1F1F")
DATA_FONT_NUM = Font(name="Calibri", size=11, color="1F1F1F")
DATA_FONT_BOLD = Font(name="Calibri", bold=True, size=11, color="1F1F1F")
DATA_ALIGN_TEXT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN_NUM = Alignment(horizontal="right", vertical="center")
DATA_ALIGN_FIRST = Alignment(horizontal="left", vertical="center", indent=1)   # 客户名首列

# Section 斑马纹 (偶数/奇数交替)
SECTION_FILL_EVEN = PatternFill("solid", fgColor="F0F0F0")   # 浅灰(偶数 section)
SECTION_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")     # 纯白(奇数 section)
SECTION_FILL_NONE = PatternFill(fill_type=None)

# 客户名首列底色
CLIENT_FILL = PatternFill("solid", fgColor="EAEFF5")                # 浅蓝灰

# 边框
BORDER_SIDE = Side(style="thin", color="BDBDBF")
BORDER_THICK = Side(style="medium", color="1F3864")
BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
TOTAL_BORDER = Border(left=BORDER_THICK, right=BORDER_THICK, top=BORDER_THICK, bottom=BORDER_THICK)

# ------------------- 数据列底色调色 -------------------
# 4 色强对比循环: 蓝 / 绿 / 黄 / 橙 (每 4 列换色)
# 饱和度 0.45 (中等), 亮度 0.92 (淡,数字清晰可读)
def _build_column_fills(n: int = 24) -> list[PatternFill]:
    fills: list[PatternFill] = []
    import colorsys
    palette = [
        (0.60, 0.92, 0.45),   # 蓝色  (D1-D4)
        (0.32, 0.92, 0.45),   # 绿色  (D5-D8)
        (0.14, 0.92, 0.45),   # 黄色  (D9-D12)
        (0.06, 0.92, 0.45),   # 橙色  (D13-D16)
        (0.60, 0.87, 0.45),   # 蓝色2 (D17-D20)
        (0.32, 0.87, 0.45),   # 绿色2 (D21-D22 + A组)
        (0.14, 0.87, 0.45),   # 黄色2
        (0.06, 0.87, 0.45),   # 橙色2
    ]
    for i in range(n):
        h, l, s = palette[i % len(palette)]
        r, g, b = colorsys.hls_to_rgb(h, l, s)
        hexc = "{:02X}{:02X}{:02X}".format(int(r * 255), int(g * 255), int(b * 255))
        fills.append(PatternFill("solid", fgColor=hexc))
    return fills


COLUMN_FILLS = _build_column_fills(24)
COLUMN_FONT = Font(name="Calibri", size=11, color="1F1F1F")


def _is_data_column_name(header_name: str) -> bool:
    """判断表头名是否属于"应上底色的数据列"。
    数据列的语义:每行有非零/非空数字(D1..D22 / A / A1..A12 / 1级 / 2级 / 年份)

    注意:**换算率、备注**这些"辅助信息列"**不上底色**——它们要么是单个数字、要么是文字评论,
    视觉上和数据矩阵混在一起反而混乱(而且我们最终也不会输出这些列,所以更不需要识别)。
    """
    if header_name is None:
        return False
    n = _normalize_header(header_name)
    if not n:
        return False
    # 年份列:应用数据列样式(右对齐+底色)
    if n == "年份":
        return True
    # 1 级 / 2 级
    if n in ("1级", "2级", "一级", "二级", "等级1", "等级2"):
        return True
    # D1..D22
    if len(n) >= 2 and n[0] == "d" and n[1:].isdigit():
        return True
    # A / A1..A12  (排除 "辅助列" 等别名,因为 _normalize 已去空格,纯 "a" 才命中)
    if n == "a" or (n.startswith("a") and n[1:].isdigit()):
        return True
    return False


def _is_color_header(header_name: str) -> bool:
    """判断表头名是否为「色号列」(D1..D29 / A / A1..A14 / A31)。

    色号列不在预剔除空列时删除 —— 它们是合并后按列对齐的关键槽位,
    即使某子表整列为空也要保留到合并后、由整组空列扫描统一清理。
    """
    if header_name is None:
        return False
    n = _normalize_header(header_name)
    if not n:
        return False
    # D1..D29
    if len(n) >= 2 and n[0] == "d" and n[1:].isdigit():
        return True
    # A / A1..A14 / A31
    if n == "a" or (n.startswith("a") and n[1:].isdigit()):
        return True
    return False


def _compute_used_data_cols(
    ws, sections: list[dict], header_rows: set[int], cm: ColumnMap | None
) -> set[int]:
    """返回「数据列」中实际有非空数据的 Excel 列号(1-based)集合。

    改为按**表头名字**判定谁是数据列 (vs 之前按">=7 + helper_idx" 位置):
    - 凡表头匹配 _is_data_column_name:1级 / 2级 / D1..D22 / A / A1..A12 / 换算率 / 备注 ...
    - 凡表头匹配 cm.helper_idx / cm.nosort_idx / 型号 / 类型 / 客户名:不上色
    - 这种按名字匹配方式天然兼容源表列位置的随意挪动

    空列不参与上色,严格满足「空行空列不处理」需求。
    """
    # 找第一个 header 行,解析哪些 ws 列号是"数据列"
    if not header_rows:
        return set()
    first_header = min(header_rows)
    data_cols_ws: set[int] = set()
    # 排除集合 (ws 列号 1-based)
    excluded: set[int] = set()
    if cm is not None:
        # ws 列号 = cm.col_idx + 1 - (cm.col_idx 之前有几个被丢的列)
        # 简化:遍历第一行,逐个判断名字,排除被识别的列名
        pass
    for c in range(1, ws.max_column + 1):
        name = ws.cell(first_header, c).value
        if _is_data_column_name(name):
            data_cols_ws.add(c)
    # 排除特殊列:模型 / 类型 — 它们在 ws 中已重映射到不同列号
    # 但在我们的实现里,我们**已经在 _parse_column_map** 阶段标记了 model / type 列
    # helper / nosort 列已从 ws 中删除,所以 ws 中已不存在
    # 因此 data_cols_ws 已经天然排除了 helper/nosort(已被 _strip 删除)
    #     也天然排除了客户名(它的表头是 None / 空字符串, _is_data_column_name 返回 False)
    # 只需要再排除: cm.model_idx / cm.type_idx / cm.year_idx 对应的 ws 列号
    # 注意: year_idx 虽然应用数据列样式,但它本身是业务维度字段(像型号/类型一样),
    #       不应在这里被当作"需要过滤空值的数据列"来处理
    for orig_idx in (cm.model_idx if cm else None, cm.type_idx if cm else None, cm.year_idx if cm else None):
        if orig_idx is None or orig_idx < 0:
            continue
        # 翻译 orig_idx → ws col:假设每张子表的 strip 不变(在 _apply_styles 调用前都是同结构),
        # 用第一个 section 的数据行做样本是不行的。
        # 我们改用遍历第一行(表头行),看哪个 ws 列号对应 model/type/year 的名字
        # 但更简单:遍历 ws 的所有列,匹配名字
        for c in range(1, ws.max_column + 1):
            n = ws.cell(first_header, c).value
            if n is None:
                continue
            target = (
                "型号" if orig_idx == cm.model_idx 
                else "类型" if orig_idx == cm.type_idx 
                else "年份" if orig_idx == cm.year_idx 
                else ""
            ) if cm else ""
            if target and _normalize_header(n) == _normalize_header(target):
                excluded.add(c)
                break
    data_cols_ws -= excluded

    # 过滤"实际有数据"的列
    used: set[int] = set()
    for sec in sections:
        for r in range(sec["data_first_row"], sec["data_last_row"] + 1):
            if r in header_rows:
                continue
            if _is_total_row_in_ws(ws, r, cm):
                continue
            for c in data_cols_ws:
                v = ws.cell(r, c).value
                if v is None:
                    continue
                if isinstance(v, str) and not v.strip():
                    continue
                used.add(c)
    return used


def _is_total_row_in_ws(ws, row_idx_1based: int, cm: ColumnMap | None) -> bool:
    """判断已写到 ws 的某行是否为合计行。用 column_map 决定「型号」所在 Excel 列。"""
    model_col_1based: int | None = None
    if cm is not None and cm.model_idx is not None:
        # 模型:helper_idx 通常是原表的最后一列(原 N 列),输出 ws 只保留到第 helper_idx 列。
        # 但「型号 / 类型」是写到 helper_idx 之前的位置,所以 col = cm.model_idx + 1
        model_col_1based = cm.model_idx + 1
    else:
        model_col_1based = 2  # 兜底
    v = ws.cell(row_idx_1based, model_col_1based).value
    return bool(v) and isinstance(v, str) and v.strip() == "合计"


# ------------------- 数据读取 & 过滤 -------------------
def _is_blank_row(row: list) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


def _is_blank_cell(v, treat_zero_as_blank: bool = False) -> bool:
    """单元格是否视为「空」:None、空字符串、全空白字符串。
    treat_zero_as_blank=True 时,数值 0 也视为空(用于预剔除阶段清理纯 0 占位列)。

    默认 False 是为了不破坏"业务上的 0 值"(0 库存/0 数量)被保留;
    但当整列都是 0 时多半是占位列,可以剔除。
    """
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    if treat_zero_as_blank and isinstance(v, (int, float)) and v == 0:
        return True
    return False

def _read_sub_file_rows(path: Path, log: list[str] | None = None) -> tuple[list[list], ColumnMap]:
    """读取子表全部行 + 解析表头列映射。

    返回 (rows, column_map):
    - rows: 跳过完全空行的二维数组，列数=该子表的**真实最大列数**(保留尾部真实空 cell)
      不再做「补齐/截断到固定 num_cols」；后续由调用方根据组内所有子表的 max 决定输出列数。
    - column_map: 「型号」「类型」「辅助列」的位置(基于表头名字识别,源表列顺序可变化)

    log: 可选,接收 [去空行] 之类的诊断日志(写入调用方的日志列表)
    """
    if log is None:
        log = []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    raw: list[list] = []
    header: list | None = None
    for row in ws.iter_rows(values_only=True):
        row_list = list(row)
        # 不再补齐/截断到固定列数;保留源表的真实宽度。
        if _is_blank_row(row_list):
            continue
        if header is None:
            header = row_list
        raw.append(row_list)
    cm = _parse_column_map(header) if header else ColumnMap()
    # ===== 移除「型号列空」的伪数据行 =====
    # 拆分脚本有时会在每张子表末尾(或段间)留下"型号列空、但 1级/2级 有 0"的占位行
    # (合计行的前一行/段尾回车)。这些不是真实型号 → 应剔除。
    # 仅当型号列存在且唯一时执行。
    if cm.model_idx is not None and raw and len(raw) >= 2:
        before = len(raw) - 1   # 减表头
        kept: list[list] = [raw[0]]
        for r in raw[1:]:
            if cm.model_idx >= len(r):
                continue
            v = r[cm.model_idx]
            # 型号列空:None / 空字符串 / 0 都算空(型号必须是字符串,数字 0 是无效占位)
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            if isinstance(v, (int, float)) and v == 0:
                continue
            kept.append(r)
        after = len(kept) - 1
        if before != after:
            log.append(
                f"  [去空行-型号] {path.name}: 移除 {before - after} 个型号列空行 "
                f"({before} → {after} 数据行)"
            )
        raw = kept
    return raw, cm


# 用户明确要求:**无论数据是否为空,以下列名一律不展示**(始终丢弃)
ALWAYS_DROP_HEADER_ALIASES: tuple[str, ...] = ("辅助列", "辅助2", "换算率", "备注")


def _always_drop_indices(header: list) -> list[int]:
    """根据表头名字返回需要「始终丢弃」的列下标(0-based)。

    规则:遍历 header,凡是名字归一化后命中 ALWAYS_DROP_HEADER_ALIASES 任一项的列都丢弃。
    用于在合并前先把"用户明确不展示"的列从每张子表中剔除,避免它们污染最大列宽计算
    / 干扰后续空列扫描 / 让最终表格出现不希望展示的列。
    """
    if not header:
        return []
    targets = {_normalize_header(a) for a in ALWAYS_DROP_HEADER_ALIASES}
    return [i for i, name in enumerate(header) if _normalize_header(name) in targets]


def _compute_max_cols(per_sub_rows: list[list[list]]) -> int:
    """根据组内所有子表的真实宽度,返回目标列数(所有子表的最大行宽)。

    设计原则:不让人手动指定列数;以拆分数据本身的实际宽度为准。
    若全为空表(没有任何行),兜底返回 0。
    """
    max_w = 0
    for rows in per_sub_rows:
        for r in rows:
            if len(r) > max_w:
                max_w = len(r)
    return max_w


def _drop_blank_columns(
    header: list, rows: list[list], treat_zero_as_blank: bool = False
) -> tuple[list, list[list], list[int]]:
    """从「表头 + 数据行」中删除「全空列」,并返回剔除后的表头、数据、所有被删除的列索引(原 0-based)。

    判定:对每个列下标 j,若 rows[*][j] 在**所有数据行**中都为空,则该列视为「全空列」并移除
    (表头是否有值不影响判定 —— 表头为空但数据有值要保留;表头有值但数据全空也要移除,
    保留空表头会让表头出现"空洞",不符合展示习惯)。
    treat_zero_as_blank: 为 True 时把数值 0 也视为空(用于清理纯占位 0 列)。

    返回:
        new_header: 压缩后的表头
        new_rows:   压缩后的所有数据行(每行已被裁短到 new_header 长度)
        removed_idx: 被删除的列在原表中的 0-based 下标(升序)
    """
    if not header and not rows:
        return header, rows, []
    width = max(
        len(header),
        max((len(r) for r in rows), default=0),
    )
    # 1. 判定哪些列要删:数据行该列全部为空(表头不影响)
    blank_cols: set[int] = set()
    for j in range(width):
        col_all_blank = True
        for r in rows:
            if j < len(r) and not _is_blank_cell(r[j], treat_zero_as_blank=treat_zero_as_blank):
                col_all_blank = False
                break
        if col_all_blank:
            blank_cols.add(j)
    if not blank_cols:
        return header, rows, []
    # 2. 重建压缩后的 header + rows
    keep = [j for j in range(width) if j not in blank_cols]
    new_header = [header[j] if j < len(header) else None for j in keep]
    new_rows: list[list] = []
    for r in rows:
        new_rows.append([r[j] if j < len(r) else None for j in keep])
    return new_header, new_rows, sorted(blank_cols)


def _compress_blank_color_cols(path: Path) -> None:
    """对复制后的子表文件原地压缩「色号列全空」的列。

    仅针对色号列(D1..D29 / A / A1..A14 / A31)做压缩,
    其他列(如客户组/型号/类型/1级/2级/辅助列)保留不动。
    只删除「整列数据行全为 None/0」且「表头属于色号列」的列。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return
    header = list(rows[0])
    data = rows[1:]
    width = max(len(header), max((len(r) for r in data), default=0))
    blank_color_cols: set[int] = set()
    for j in range(width):
        if not _is_color_header(header[j] if j < len(header) else None):
            continue
        if all(
            j >= len(r) or _is_blank_cell(r[j], treat_zero_as_blank=True)
            for r in data
        ):
            blank_color_cols.add(j)
    if not blank_color_cols:
        wb.close()
        return
    keep = [j for j in range(width) if j not in blank_color_cols]
    new_header = [header[j] if j < len(header) else None for j in keep]
    new_rows = [[r[j] if j < len(r) else None for j in keep] for r in data]
    ws.delete_rows(1, ws.max_row)
    ws.append(new_header)
    for r in new_rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def _remap_column_map(cm: ColumnMap, removed: list[int]) -> ColumnMap:
    """对 ColumnMap 的列索引做「压缩后重映射」。

    - removed: 已删除列的 0-based 索引(升序)
    - 索引 >= 0 的字段(model/type/helper/nosort)若被删 → 置 None / 默认值
    - 其余字段按"被删列数偏移量"前移
    """
    if not removed:
        return cm
    return _remap_column_map_with_drop(cm, removed)


def _remap_column_map_with_drop(cm: ColumnMap, drop: list[int]) -> ColumnMap:
    """根据"被丢弃列下标集合 drop"(0-based,升序)对 ColumnMap 的列索引做重映射。

    与 _remap_column_map 等价 —— 拆分命名仅为了在按"白名单列名预剔除"场景下语义更清晰。
    """
    if not drop:
        return cm
    removed_set = set(drop)
    def _shift(idx: int | None) -> int | None:
        if idx is None or idx < 0:
            return idx
        if idx in removed_set:
            return None
        return idx - sum(1 for x in drop if x < idx)

    new_cm = ColumnMap(
        model_idx=_shift(cm.model_idx),
        type_idx=_shift(cm.type_idx),
        year_idx=_shift(cm.year_idx),
        helper_idx=_shift(cm.helper_idx) if cm.helper_idx not in removed_set else DEFAULT_HELPER_COL_IDX,
        nosort_idx=_shift(cm.nosort_idx),
        max_used_col_idx=max(0, cm.max_used_col_idx - len(drop)),
    )
    return new_cm


def _strip_columns(row: list, drop_idx: list[int]) -> list:
    """从单行中按 drop_idx(0-based,升序)丢弃列,返回新行。

    行长度可能 < max(drop_idx),不在范围内的下标自动跳过(不会越界)。
    """
    if not drop_idx:
        return row
    drop_set = set(drop_idx)
    return [v for j, v in enumerate(row) if j not in drop_set]


# ------------------- 表头列名识别 (容错:源表列位置可变化) -------------------
# 源表格中可能存在这些列(其它列保持原顺序输出):
#   - 型号    :用于识别「合计行」,主排序键 2
#   - 类型    :主排序键 1
#   - 辅助列  :最后一列(可能命名「显示」「辅助」「是否显示」「隐藏」等),决定行是否被丢弃
# 无论你把这些列挪到 A/B/C/...,只要表头名字匹配,脚本都能工作。
COL_NAME_MODEL = "型号"
COL_NAME_TYPE = "类型"
# 辅助列的表头可能有多种写法(逐项匹配)
COL_NAME_HELPER_ALIASES = ("显示", "辅助", "是否显示", "显示列", "显示控制", "隐藏")
# 「辅助列 / 辅助」是过滤控制(显示/不显示); 「辅助2」是排序控制(不排序 / 空)
COL_NAME_HELPER_FILTER_ALIASES = ("显示", "是否显示", "显示列", "显示控制", "隐藏", "辅助列", "辅助")
COL_NAME_HELPER_SORT_ALIASES = ("辅助2", "不排序", "不排序列", "排序控制", "辅助 2", "aux2")
DEFAULT_HELPER_COL_IDX = -1   # 默认仍是最后一列(向后兼容旧表)
DEFAULT_NOSORT_COL_IDX = -1   # 默认不存在该列


@dataclass
class ColumnMap:
    """从一张子表的表头行解析出的列映射。所有索引都是 0-based 数组下标,
    适用于 _read_sub_file_rows 返回的 row 列表。

    关键字段:
    - model_idx:  「型号」所在列(用于合计行识别 + 排序次键)
    - type_idx:   「类型」所在列(用于排序主键)
    - year_idx:   「年份」所在列(新增,不参与排序,应用数据列样式)
    - helper_idx: 过滤控制列(显示 / 不显示; 不显示的行被丢)
    - nosort_idx: 排序控制列(不排序 / 空; 标了"不排序"的行不参与排序,保留原顺序)
    - max_used_col_idx: 表头最大下标(用于数据列染色范围判定)
    """
    model_idx: int | None = None   # 「型号」所在列
    type_idx: int | None = None    # 「类型」所在列
    year_idx: int | None = None    # 「年份」所在列
    helper_idx: int = DEFAULT_HELPER_COL_IDX   # 过滤控制列(默认最后一列)
    nosort_idx: int | None = None  # 排序控制列(默认不存在 → 全量排序)
    max_used_col_idx: int = 0


def _normalize_header(v) -> str:
    """表头名归一化:去空白、统一小写、去掉所有空格与全角空格,用于模糊匹配。"""
    if v is None:
        return ""
    s = str(v).strip().casefold()
    return s.replace(" ", "").replace("\u3000", "")


def _parse_column_map(header: list) -> ColumnMap:
    """从表头行构建 ColumnMap。

    字段识别规则:
    - 「型号」「类型」「年份」:精确匹配(casefold)
    - helper(filter): 「显示 / 辅助 / 是否显示 / 显示列 / 显示控制 / 隐藏 / 辅助列」任一别名
                      — 取**最后**一个匹配列,通常对应源表那个"最后显示控制列"
    - nosort:        「辅助2 / 不排序列 / 排序控制 / 辅助 2 / aux2」任一别名

    找不到对应列时:
    - 型号/类型/年份 → None(兜底退化)
    - helper → -1(原约定:辅助视为末列,自动被丢)
    - nosort → None(不启用"不排序"功能,全部数据行参与排序)
    """
    cm = ColumnMap()
    for i, raw in enumerate(header):
        name = _normalize_header(raw)
        if name == "型号" and cm.model_idx is None:
            cm.model_idx = i
        elif name == "类型" and cm.type_idx is None:
            cm.type_idx = i
        elif name == "年份" and cm.year_idx is None:
            cm.year_idx = i
        else:
            # 别名匹配:按优先级顺序遍历,后写的覆盖前面的(谁靠后谁赢,符合用户把"辅助列"挪到中间的场景)
            filter_match = any(_normalize_header(a) == name for a in COL_NAME_HELPER_FILTER_ALIASES)
            nosort_match = any(_normalize_header(a) == name for a in COL_NAME_HELPER_SORT_ALIASES)
            if filter_match:
                cm.helper_idx = i
            if nosort_match:
                cm.nosort_idx = i
    cm.max_used_col_idx = len(header) - 1 if header else 0
    return cm


def _is_total_row(row: list, cm: ColumnMap) -> bool:
    """判定某行是否为合计行:型号列 == '合计'。
    型号列可能已被挪到任意位置(通过 column_map 解析);找不到则按 False。
    """
    idx = cm.model_idx if cm is not None else None
    if idx is None or idx >= len(row):
        return False
    v = row[idx]
    return isinstance(v, str) and v.strip() == "合计"


def _fallback_total_idx(rows: list[list], cm: ColumnMap) -> int | None:
    """兜底:没有任何型号=='合计'的行时,取末尾倒数第二个非空数据行作为合计。"""
    if len(rows) < 2:
        return None
    return len(rows) - 2


def _filter_rows(
    rows: list[list], cm: ColumnMap
) -> tuple[list[list], list[list], list[list], int, int]:
    """返回 (header_row, data_rows, total_rows, drop_count, total_count)。

    - header_row: 第一行(子表的列名)单独保留
    - total_rows: 第一个型号=='合计'的行;兜底用最后第二个非空数据行
    - data_rows: 其余数据行
    - 辅助列(ColumnMap.helper_idx)== '不显示' 且不是合计行 → 丢弃
    - 辅助列未匹配时回退到最后一列
    """
    total_idx = None
    for i, r in enumerate(rows):
        if i == 0:
            continue
        if _is_total_row(r, cm):
            total_idx = i
            break
    if total_idx is None:
        total_idx = _fallback_total_idx(rows, cm)

    helper_idx_raw = cm.helper_idx if cm is not None else -1
    drop = 0
    header: list[list] | None = None
    totals: list[list] = []
    data: list[list] = []
    for i, r in enumerate(rows):
        if i == 0:
            header = r
            continue
        if i == total_idx:
            totals.append(r)
            continue
        helper = None
        if helper_idx_raw is not None and len(r) > 0:
            resolved = helper_idx_raw if helper_idx_raw >= 0 else len(r) + helper_idx_raw
            if 0 <= resolved < len(r):
                helper = r[resolved]
        if isinstance(helper, str) and helper.strip() == "不显示":
            drop += 1
            continue
        data.append(r)
    return header or [], data, totals, drop, len(rows)


# ------------------- 数据排序 -------------------
def _sort_key_for_data_row(row: list, cm: ColumnMap):
    """生成数据行的排序键:(类型, 型号)。

    - 优先用 cm.type_idx / cm.model_idx
    - 表头中找不到对应列时,退化到旧位置 row[2] / row[1]
    - None 排到末尾
    - 中文优先按拼音排序;英文/数字按字母顺序;混排时中文排在数字/字母之后
    """
    def _locale_key(v):
        if v is None:
            return (1, "")
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return (1, "")
            try:
                import locale
                return (0, s.lower())
            except Exception:
                return (0, s.casefold())
        return (0, str(v))

    if cm is not None:
        type_idx = cm.type_idx if cm.type_idx is not None else 2
        model_idx = cm.model_idx if cm.model_idx is not None else 1
    else:
        type_idx, model_idx = 2, 1
    a = _locale_key(row[type_idx]) if type_idx < len(row) else (1, "")
    b = _locale_key(row[model_idx]) if model_idx < len(row) else (1, "")
    return (a, b)


def _is_nosort_row(row: list, cm: ColumnMap) -> bool:
    """判断数据行是否被 nosort 列(辅助2 / 不排序列)标记为「不排序」。

    仅当 cm.nosort_idx 存在(>= 0)且行长度足够时才检查;否则返回 False(全部参与排序)。
    """
    if cm is None or cm.nosort_idx is None or cm.nosort_idx < 0:
        return False
    if cm.nosort_idx >= len(row):
        return False
    v = row[cm.nosort_idx]
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return False
        return s in ("不排序", "否", "N", "n", "No", "NO", "no", "0", "false", "False")
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v == 0
    return False


def _sort_data_rows(rows: list[list], cm: ColumnMap) -> list[list]:
    """对数据行按 类型 → 型号 升序排序,保留「不排序」标记行的原顺序。

    排序策略:
    1. 把数据行拆成 pinned(标了不排序,按原顺序)+ sortable(未标,正常排序)
    2. sortable 按 (类型, 型号) 排序后,接在 pinned 之后
       (pinned 在前:用户标"不排序"通常想让这些行优先展示,例如特殊型号置顶)
    3. 在稳定排序下,sortable 内同类型同型号仍保持原顺序

    没 cm.nosort_idx 时退化为纯排序(向后兼容)。
    """
    if cm is None or cm.nosort_idx is None:
        return sorted(rows, key=lambda r: _sort_key_for_data_row(r, cm))
    pinned, sortable = [], []
    for r in rows:
        if _is_nosort_row(r, cm):
            pinned.append(r)
        else:
            sortable.append(r)
    sortable_sorted = sorted(sortable, key=lambda r: _sort_key_for_data_row(r, cm))
    return pinned + sortable_sorted


# ------------------- 报表样式应用 -------------------
def _visual_width(s: str) -> float:
    """估算单元格内容在 Excel 列宽单位下的视觉宽度。

    - CJK(中日韩)统一按 2 个单位计(Excel 中英字约 1 字符宽 = 7 px,中文约 14 px)
    - 数字/英文/常见 ASCII 按 1 个单位计
    - 全角符号(·、—、… 等)按 2 计
    """
    if not s:
        return 0.0
    w = 0.0
    for ch in s:
        code = ord(ch)
        # CJK 基本区 + 扩展 A-F + 标点 + 全角符号
        if (
            0x1100 <= code <= 0x115F          # Hangul Jamo
            or 0x2E80 <= code <= 0x303E       # CJK 标点 / 部首
            or 0x3041 <= code <= 0x33FF       # 平假名 / 片假名 / CJK 符号
            or 0x3400 <= code <= 0x4DBF       # CJK 扩展 A
            or 0x4E00 <= code <= 0x9FFF       # CJK 基本
            or 0xA000 <= code <= 0xA4CF       # 彝文
            or 0xAC00 <= code <= 0xD7A3       # 韩文音节
            or 0xF900 <= code <= 0xFAFF       # CJK 兼容
            or 0xFE30 <= code <= 0xFE4F       # CJK 兼容形式
            or 0xFF00 <= code <= 0xFF60       # 全角 ASCII / 全角标点
            or 0xFFE0 <= code <= 0xFFE6       # 全角符号
        ):
            w += 2.0
        else:
            w += 1.0
    return w


def _autosize_columns(ws, min_w: float = 8.0, max_w: float = 30.0, extra_for_total: float = 4.0) -> None:
    """按列内容自适应宽度(用每列最长 cell 的真实视觉宽度 + padding)。

    设计原则:
    - **必须**用 max(不是 P95):少数长型号(如 `IN12P004GL60120`)若被截,客户读不出来
    - 默认 padding = 2,空出 1 字符边距
    - 上限 max_w = 30(防止极长 token 拉爆整张表;超长用截断或换行兜底)
    - 下限 min_w = 8(数字列 "1234567" 也至少要装下)
    - CJK 字符按 2 个单位计(用 _visual_width)
    - 合计行数字列额外 + extra_for_total,避免大数被截成 ###
    """
    from openpyxl.utils import get_column_letter
    # 先找出合计行所在的列(用于判断是否额外加宽)
    total_row_indices = set()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "合计":
                total_row_indices.add(cell.row)
                break

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0.0
        has_total = False
        for cell in col_cells:
            if cell.value is None:
                continue
            s = str(cell.value)
            max_len = max(max_len, _visual_width(s))
            if cell.row in total_row_indices:
                has_total = True
        # padding = 2 (左边距 1 + 右边距 1)
        target = max_len + 2
        # 合计行数字列额外留空间
        if has_total:
            target += extra_for_total
        # 软上限:若 target > max_w, 仍给到 max_w;超过部分会被显示为溢出
        width = max(min_w, min(max_w, target))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _is_numeric_header(name: str) -> bool:
    """判断表头是否代表"数字列"(用于决定右对齐 + 数字格式)。
    1级 / 2级 / D1..D22 / A / A1..A12 / 年份 都是数字列。
    """
    if name is None:
        return False
    n = _normalize_header(name)
    if not n:
        return False
    if n == "年份":
        return True
    if n in ("1级", "2级", "一级", "二级", "等级1", "等级2"):
        return True
    if len(n) >= 2 and n[0] == "d" and n[1:].isdigit():
        return True
    if n == "a" or (n.startswith("a") and n[1:].isdigit()):
        return True
    return False


def _apply_styles(ws, sections: list[dict], cm: ColumnMap | None = None) -> None:
    """应用专业报表样式 (client-facing preset)。

    视觉系统:
    1. 表头行:深海军蓝底 + 白色粗体 + 居中 + 行高 32 + wrap_text
    2. 数据行:section 斑马纹 (奇浅 / 偶白)
       - 文本列(型号/类型):居中 + 允许换行 (避免超长型号挤出)
       - 数字列(1级/2级/D..A):右对齐 + 千分位格式
       - 客户名首列(A):浅蓝灰底 + 左对齐 + 缩进 1
    3. 合计行:橙黄底 + 粗体 + 顶边加粗 + 全行右对齐(数字)/居中(文本)
    4. 数据列底色循环 (蓝族 12 + 紫族 12),帮助识别列归属
    5. 全表细线边框 (D9D9D9),不刺眼
    6. 冻结首行 + 冻结首列 (A 列客户名)

    cm (ColumnMap) 用于判定「合计行」(型号列位置由 cm 决定)。
    """
    n_cols = ws.max_column
    n_rows = ws.max_row
    header_rows = {sec["header_row"] for sec in sections}

    # 把每个数据列标上序号 + 数据列属性 (用于背景色 + 数字格式)
    # 用 _is_data_column_name 找数据列
    data_col_ordered: list[tuple[int, str]] = []   # (ws 列号 1-based, header 名)
    for c in range(1, n_cols + 1):
        name = ws.cell(1, c).value
        if _is_data_column_name(name):
            data_col_ordered.append((c, str(name) if name else ""))

    # ===== 1) 全表边框 =====
    for row in ws.iter_rows(min_row=1, max_row=n_rows, max_col=n_cols):
        for cell in row:
            cell.border = BORDER

    # ===== 2) 表头行 =====
    for sec in sections:
        hr = sec["header_row"]
        for cell in ws[hr]:
            cell.fill = HEAD_FILL
            cell.font = HEAD_FONT
            cell.alignment = HEAD_ALIGN
        ws.row_dimensions[hr].height = 36

    # ===== 3) 数据行染色 + 对齐 =====
    # 对每个 section 内部按行号判断"奇偶行"做斑马纹 (section 内重置计数,跨 section 重新开始)
    # section 整体偶数/奇数 (在分组中的位置) 也切换底色,但为了客户视觉清爽,
    # 这里采用"全表统一白底 + 数据列底色循环",不去做斑马纹(避免和数据列底色冲突)
    for sec_idx, sec in enumerate(sections):
        for r in range(sec["data_first_row"], sec["data_last_row"] + 1):
            is_total = _is_total_row_in_ws(ws, r, cm)
            for cell in ws[r]:
                col = cell.column
                cell.border = BORDER if not is_total else TOTAL_BORDER
                if is_total:
                    # 合计行
                    cell.fill = TOTAL_FILL
                    cell.font = TOTAL_FONT
                    if col == 1:
                        cell.alignment = DATA_ALIGN_FIRST
                    else:
                        # 数字列右对齐,文本列居中
                        name = ws.cell(sec["header_row"], col).value if sec["header_row"] <= ws.max_row else None
                        if _is_numeric_header(name):
                            cell.alignment = DATA_ALIGN_NUM
                            cell.number_format = '#,##0;-#,##0;""'
                        else:
                            cell.alignment = TOTAL_ALIGN
                    continue

                # 数据行
                # section 斑马纹: 偶数 section 浅灰底, 奇数 section 白底
                sec_zebra = SECTION_FILL_EVEN if sec_idx % 2 == 0 else SECTION_FILL_ODD
                if col == 1:
                    # 客户名首列: 蓝色浅底 + 粗体
                    cell.fill = CLIENT_FILL
                    cell.font = DATA_FONT_BOLD
                    cell.alignment = DATA_ALIGN_FIRST
                else:
                    # 检查是否在数据列内
                    data_idx = next((i for i, (wc, _) in enumerate(data_col_ordered) if wc == col), None)
                    if data_idx is not None and cell.value is not None and not (isinstance(cell.value, str) and not cell.value.strip()):
                        # 数据列有值:上列底色 + 数字格式
                        fill = COLUMN_FILLS[data_idx % len(COLUMN_FILLS)]
                        cell.fill = fill
                        cell.font = COLUMN_FONT
                        cell.alignment = DATA_ALIGN_NUM
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0;-#,##0;""'
                        elif isinstance(cell.value, str):
                            try:
                                v = float(cell.value)
                                cell.value = v
                                cell.number_format = '#,##0;-#,##0;""'
                                cell.alignment = DATA_ALIGN_NUM
                            except (ValueError, TypeError):
                                cell.alignment = DATA_ALIGN_TEXT
                    else:
                        # 非数据列 / 空格: section 斑马纹底色
                        cell.fill = sec_zebra
                        cell.font = DATA_FONT
                        cell.alignment = DATA_ALIGN_TEXT

    # ===== 4) 行高 =====
    for r in range(2, n_rows + 1):
        if r in header_rows:
            continue
        ws.row_dimensions[r].height = 22

    # ===== 5) 冻结首行 + 客户列 =====
    ws.freeze_panes = "B2"

    # ===== 6) 自适应列宽 =====
    _autosize_columns(ws)


# ------------------- 合并主函数 -------------------
def _build_merged_file(
    group: dict,
    available: set[str],
    src_dir: Path,
    output_dir: Path,
    log: list[str],
) -> tuple[Path, int, int] | None:
    """返回 (path, rows_kept, rows_dropped). 仍然无数据时返回 None.

    拆分列数逻辑(2026-08 优化):
    1) 先按表头名始终丢弃:辅助列 / 辅助2 / 换算率 / 备注 — 这些列用户明确不要展示,
       不论数据是否为空一律从每张子表中剔除(同时调整 ColumnMap 索引)
    2) 列数自动确定:扫描组内所有子表(含已剔除辅助列)的真实最大列数,作为合并目标列数
    3) 空列自动移除:对整组合并结果(含所有子表的表头 + 数据 + 合计)做"数据行全空列"扫描,
       命中后整列剔除(同时更新 ColumnMap 的索引),压缩列数优化展示
    """
    out_name = group["a_col"]
    if not out_name:
        return None

    group_dir = output_dir / out_name
    group_dir.mkdir(parents=True, exist_ok=True)

    merged_path = group_dir / f"{out_name}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = out_name[:31]

    # 默认 Workbook 已含 1 行；删除它并把指针归零，避免 header_row 不等于 1
    ws.delete_rows(1)
    current_row = 0

    sections: list[dict] = []
    rows_kept_total = 0
    rows_dropped_total = 0
    # 全局 cm:以第一张有效子表为准,后续写出 / 排序 / 染色均按它进行。
    # (分类表的源表通常用同一模板,但即使不同也会被合并到同一 ws,只要头部一致即可)
    global_cm: ColumnMap | None = None

    for sub in group["sub_files"]:
        candidate = src_dir / f"{sub}.xlsx"
        if sub not in available or not candidate.exists():
            log.append(f"  [跳过-缺文件] {sub}")
            continue
        raw_rows, cm = _read_sub_file_rows(candidate, log)

        # ===== 子表内「色号列全空」伪空行剔除 =====
        # 规则:色号列(D1..D22 / A / A1..A12 / A31 等,即 _is_data_column_name 命中但
        # 不含 1级/2级)中,所有列都为空 → 该行算伪空行,剔除。
        # 1级/2级 有 0 不算"色号列"—— 1级/2级 是库存等级,与具体色号无关。
        # 重要:必须在「预剔除空列」之前做 —— 否则像 Y客户2级这种"整张表色号列全空"
        # 的子表会被预剔除把 1级/2级+色号列整列删掉,这里 header 已无 D1..A14
        # 等槽位 → 找不到 color_indices → 永不剔除。
        if raw_rows and len(raw_rows) >= 2 and cm.model_idx is not None:
            sub_header = raw_rows[0]
            color_indices = [
                i for i, name in enumerate(sub_header)
                if _is_data_column_name(name)
                and _normalize_header(name) not in ("1级", "2级", "一级", "二级", "等级1", "等级2")
            ]
            l1l2_indices = [
                i for i, name in enumerate(sub_header)
                if _normalize_header(name) in ("1级", "2级", "一级", "二级", "等级1", "等级2")
            ]
            base_data_indices = [
                i for i, name in enumerate(sub_header)
                if _is_data_column_name(name)
            ]
            use_l1l2 = bool(l1l2_indices)
            if color_indices:
                before = len(raw_rows) - 1
                kept_rows: list[list] = [raw_rows[0]]
                for r in raw_rows[1:]:
                    # 色号列判定:把 0/None/空字符串 都视为空(色号列里 0 = 没库存)
                    color_all_blank = True
                    for i in color_indices:
                        if i < len(r) and not _is_blank_cell(r[i], treat_zero_as_blank=True):
                            color_all_blank = False
                            break
                    if not color_all_blank:
                        kept_rows.append(r)
                        continue
                    # 色号列全空 → 进一步判定该行是否有任何库存
                    if use_l1l2:
                        has_l1l2 = False
                        for i in l1l2_indices:
                            if i < len(r):
                                v = r[i]
                                if isinstance(v, (int, float)) and v > 0:
                                    has_l1l2 = True
                                    break
                        if not has_l1l2:
                            continue  # 剔除
                    else:
                        has_any = False
                        for i in base_data_indices:
                            if i < len(r):
                                v = r[i]
                                if isinstance(v, (int, float)) and v > 0:
                                    has_any = True
                                    break
                        if not has_any:
                            continue  # 剔除
                    kept_rows.append(r)
                after = len(kept_rows) - 1
                if before != after:
                    log.append(
                        f"  [去空行-色号] {sub}: 移除 {before - after} 个色号列全空行 "
                        f"({before} → {after} 数据行)"
                    )
                raw_rows = kept_rows

        # ===== 始终丢弃用户明确不展示的列 =====
        # (辅助列 / 辅助2 / 换算率 / 备注) — 不论数据是否有值,按表头名命中即剔除。
        # 不同子表的列顺序可能不同,所以每张子表独立识别;同时把 cm 索引按偏移前移,
        # 保证"合计行判定 / 排序键"等仍指对正确位置。
        sub_header = raw_rows[0] if raw_rows else None
        sub_drop_idx = _always_drop_indices(sub_header)
        if sub_drop_idx:
            log.append(
                f"  [预剔除] {sub}: 按表头名丢弃 {len(sub_drop_idx)} 列 "
                f"(列号 {sub_drop_idx}) → 列数 {len(sub_header) - len(sub_drop_idx)}"
            )
            raw_rows = [_strip_columns(r, sub_drop_idx) for r in raw_rows]
            # cm 索引前移(指向新位置);用 sub_drop_idx 的下标集合做"左侧计数后减一"重映射
            cm = _remap_column_map_with_drop(cm, sub_drop_idx)

        # ===== 子表内「数据全空列」预剔除 =====
        # 拆分脚本有时会产生"占位列"(比如每日库存场景里的 D1-D28 / A1-A14 / A31
        # 等扩展位,大部分行该列为空)。这些空列在不同子表的位置不完全一致,
        # 导致后面"整组合并后的全空列扫描"无法剔除(不同子表错位无法全部为空)。
        # 在此按"本子表内数据行全空"先压缩一遍,可显著减少最终表格的空洞列。
        # 判定:对每列 j,只要子表数据行中存在 1 个非空 cell,该列就保留。
        # 表头是否为空不影响 —— 因为表头本身可能就没值但位置是"必须留的"语义槽。
        # 注意:色号列(D1..D29 / A / A1..A14 / A31)不在此剔除 —— 它们是合并后
        # 按列对齐的关键槽位,即使某子表整列为空也要保留到合并后、由整组空列扫描统一清理。
        sub_header_for_drop = raw_rows[0] if raw_rows else []
        if raw_rows and len(raw_rows) >= 2:
            sub_data = raw_rows[1:]
            width = max(
                len(raw_rows[0]) if raw_rows[0] else 0,
                max((len(r) for r in sub_data), default=0),
            )
            blank_cols: set[int] = set()
            for j in range(width):
                header_name = sub_header_for_drop[j] if j < len(sub_header_for_drop) else None
                if _is_color_header(header_name):
                    continue  # 跳过色号列,即使整列空也保留
                if all(
                    j >= len(r) or _is_blank_cell(r[j], treat_zero_as_blank=True)
                    for r in sub_data
                ):
                    blank_cols.add(j)
            if blank_cols:
                drop_sorted = sorted(blank_cols)
                log.append(
                    f"  [预剔除] {sub}: 数据全空列丢弃 {len(drop_sorted)} 列 "
                    f"(列号 {drop_sorted}) → 列数 {width - len(drop_sorted)}"
                )
                raw_rows = [_strip_columns(r, drop_sorted) for r in raw_rows]
                cm = _remap_column_map_with_drop(cm, drop_sorted)

        if global_cm is None:
            global_cm = cm
        header_row, data_rows, total_rows, drop, total = _filter_rows(raw_rows, cm)
        # 数据行排序:类型 → 型号 升序(表头与合计行不参与排序)
        data_rows_sorted = _sort_data_rows(data_rows, cm) if data_rows else []
        kept_rows = ([header_row] if header_row else []) + data_rows_sorted + total_rows
        log.append(
            f"  [过滤] {sub}: 保留 {len(kept_rows)} / 全部 {total}  (丢弃 {drop})"
        )

        if not kept_rows:
            continue

        # 写表:每个 section 写"数据+合计行"。
        # 表头统一只写在最前面(由后续整组空列压缩逻辑保留第 1 行作表头),
        # 避免每个子表自带表头导致中间出现重复的「客户组/型号/类型...」行
        # (这些行里的列名会撑住"本子表独有的列",妨碍后续压缩)。
        # 不在段间插空行 —— 段间视觉分隔靠首列(客户组)的颜色 + 合并的合计行天然分隔。
        # 写入策略:用 ws.max_row 跟踪实际位置,ws.append 总是写到 max_row+1;
        # 因此 current_row 应在 ws.append 之后 = ws.max_row(避免 +=1 算多 1 行)。
        is_first_section = len(sections) == 0
        if is_first_section:
            section_header_row = current_row + 1   # 第一次 append 会写到 max_row+1
            ws.append(kept_rows[0])                # 表头(只写一次)
            current_row = ws.max_row
            section_data_first = current_row + 1
        else:
            # 非首段:不写表头,直接续写数据(段间密排)
            section_header_row = sections[0]["header_row"]
            section_data_first = current_row + 1
        for r in kept_rows[1:]:
            ws.append(r)
            current_row = ws.max_row
        data_first_row = section_data_first
        data_last_row = current_row

        rows_kept_total += len(kept_rows)
        rows_dropped_total += drop
        sections.append({
            "header_row": section_header_row,
            "data_first_row": data_first_row,
            "data_last_row": data_last_row,
        })

    if rows_kept_total == 0:
        wb.close()
        log.append(f"[空组跳过] {out_name} (无任何子表匹配或全部被过滤)")
        shutil.rmtree(group_dir)
        return None

    # ===== 列数自动确定 =====
    # 1) 取整组合并表(已写出)的最大列数 = 各 section 实际占用的最大列号
    n_cols_actual = ws.max_column
    # (openpyxl 的稀疏 cell 机制保证:对任何 r/c,ws.cell(r,c).value 都会返回 None 或实际值,
    #  不需要主动把短 section 补齐)

    # 2) 收集所有 (header_row -> 表头) + 数据行,做"全空列"扫描
    #    移除空列后再回写 ws;同时把 global_cm 的列索引按已删列做偏移重映射
    header_cells: list = []
    data_cells_by_row: dict[int, list] = {}
    for sec in sections:
        hr = sec["header_row"]
        for c in range(1, n_cols_actual + 1):
            if c <= len(header_cells):
                # 已经有该列的 header(多张子表合并),取第一个非空的
                if header_cells[c - 1] is None:
                    header_cells[c - 1] = ws.cell(hr, c).value
            else:
                header_cells.append(ws.cell(hr, c).value)
        for r in range(sec["data_first_row"], sec["data_last_row"] + 1):
            data_cells_by_row[r] = [ws.cell(r, c).value for c in range(1, n_cols_actual + 1)]

    new_header, new_rows_per_row, removed_idx = _drop_blank_columns(
        header_cells,
        [data_cells_by_row[r] for r in sorted(data_cells_by_row.keys())],
        treat_zero_as_blank=True,
    )

    if removed_idx:
        log.append(
            f"  [压缩] 移除 {len(removed_idx)} 个全空列 "
            f"(列号 {removed_idx}) → 输出列数 {len(new_header)}"
        )
        # 3) 按压缩结果整体回写 ws
        #    先清空 ws 既有内容(行/样式都会被一并清掉,不影响后续 _apply_styles)
        ws.delete_rows(1, ws.max_row)
        # 写入新表头
        ws.append(new_header)
        # 写入各 section 的数据行;按原 sections 的行号范围重新映射
        # 压缩后整张表只有 1 个 header_row(=1),各 section 紧接其后依次排列
        new_sections: list[dict] = []
        running_row = 1   # header 已写入
        # 预先把 data_cells_by_row 排序后对应的压缩行准备好
        compressed_by_orig_row: dict[int, list] = {}
        compressed_iter = iter(new_rows_per_row)
        for r in sorted(data_cells_by_row.keys()):
            compressed_by_orig_row[r] = next(compressed_iter)
        for sec in sections:
            sec_first = running_row + 1
            for r in range(sec["data_first_row"], sec["data_last_row"] + 1):
                ws.append(compressed_by_orig_row[r])
                running_row += 1
            new_sections.append({
                "header_row": 1,
                "data_first_row": sec_first,
                "data_last_row": running_row,
            })
        sections = new_sections
        # 4) 重映射 global_cm 的索引(让样式判定「型号 / 类型」位置仍正确)
        if global_cm is not None:
            global_cm = _remap_column_map(global_cm, removed_idx)

    _apply_styles(ws, sections, global_cm)
    # 裁短尾部「无任何 cell」的稀疏行 —— 列压缩后 ws.max_row 可能仍指向旧值
    # (openpyxl 不会自动收缩),这里主动用 delete_rows 把 last_real_row 之后的
    # 物理行删掉,让 Excel 不显示这些空行
    last_real_row = max((s["data_last_row"] for s in sections), default=1)
    last_real_row = max(last_real_row, 1)  # 至少保留表头
    if last_real_row < ws.max_row:
        ws.delete_rows(last_real_row + 1, ws.max_row - last_real_row)
    wb.save(merged_path)
    return merged_path, rows_kept_total, rows_dropped_total


def _copy_source_files(
    group: dict,
    available: set[str],
    src_dir: Path,
    output_dir: Path,
) -> int:
    out_name = group["a_col"]
    group_dir = output_dir / out_name
    copied = 0
    for sub in group["sub_files"]:
        src = src_dir / f"{sub}.xlsx"
        dst = group_dir / f"{sub}.xlsx"
        if sub not in available or not src.exists():
            continue
        shutil.copy2(src, dst)
        copied += 1
    return copied


def load_classification(index_path: Path) -> list[dict]:
    """读取分类表，返回每行配置: {a_col, b_col, sub_files}。"""
    wb = openpyxl.load_workbook(index_path, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        a_val, b_val = row[0], row[1]
        if a_val is None and b_val is None:
            continue
        sub_files = [str(v).strip() for v in row[2:] if v is not None and str(v).strip()]
        rows.append(
            {
                "a_col": str(a_val).strip() if a_val is not None else "",
                "b_col": str(b_val).strip() if b_val is not None else "",
                "sub_files": sub_files,
            }
        )
    return rows


# ------------------- 进程入口 -------------------
ProgressCb = Callable[[int, int, str], None]


def process(cfg: dict, progress_cb: ProgressCb | None = None) -> ProcessResult:
    """主入口。cfg 需含 src_dir/index_file/output_dir。
    列数不再由 cfg 传入:由 _build_merged_file 在每组内自动按拆分数据决定,并剔除全空列。
    progress_cb(current_idx, total, message) 在每个组开始时调用一次。"""
    result = ProcessResult()

    src_dir = Path(cfg["src_dir"])
    index_file = Path(cfg["index_file"])
    output_dir = Path(cfg["output_dir"])

    if not src_dir or src_dir == Path("."):
        raise ValueError("src_dir 未配置")
    if not index_file or index_file == Path("."):
        raise ValueError("index_file 未配置")
    if not output_dir or output_dir == Path("."):
        raise ValueError("output_dir 未配置")
    if not index_file.exists():
        raise FileNotFoundError(f"索引文件不存在: {index_file}")
    if not src_dir.exists():
        raise FileNotFoundError(f"源目录不存在: {src_dir}")

    actual_output_dir = _prepare_output_dir(output_dir, allow_unsafe=bool(cfg.get("allow_unsafe_output", False)))
    result.log.append(f"[输出] 使用目录: {actual_output_dir.name}")
    output_dir = actual_output_dir
    result.actual_output_dir = actual_output_dir

    available = {f[:-5] for f in os.listdir(src_dir) if f.lower().endswith(".xlsx")}
    groups = load_classification(index_file)
    result.groups_total = len(groups)
    if not groups:
        return result

    for idx, g in enumerate(groups, start=1):
        out_name = g["a_col"]
        log = result.log
        log.append(f"=== {out_name} ===")
        if progress_cb:
            progress_cb(idx - 1, len(groups), f"正在合并: {out_name}")
        merged = _build_merged_file(g, available, src_dir, output_dir, log)
        if merged is None:
            continue
        merged_path, kept, dropped = merged
        result.rows_total += kept
        result.rows_filtered_total += dropped
        copied = _copy_source_files(g, available, src_dir, output_dir)
        result.files_copied += copied
        result.missing_files.extend(
            sub for sub in g["sub_files"]
            if sub not in available or not (src_dir / f"{sub}.xlsx").exists()
        )
        log.append(f"  合并文件: {merged_path.name}  (有效 {kept} 行，过滤 {dropped} 行)")
        log.append(f"  源子表: 已复制 {copied} / {len(g['sub_files'])} 个")
        result.groups_merged += 1
        if progress_cb:
            progress_cb(idx, len(groups), f"完成: {out_name}")

    return result


# ------------------- CLI 入口 -------------------
def _demo_data(out_root: Path) -> tuple[Path, Path, Path]:
    """在 out_root 下生成一组演示数据,返回 (src_dir, index_file, output_dir)。

    演示场景:
      组A
        subA : 客户名|型号|类型|色A|色B|色C|数量|单价|辅助列|换算率|备注
               含 3 个辅助列,且 色A/色C 数据全空
        subB : 客户名|型号|类型|数量|备注|辅助列|辅助2
               列顺序不同,且 数量/色A(空) 等与 subA 错位
               ⚠️ 注意:不同子表同一位置不一定都是空,压缩只对整组同时空才生效
    """
    src = out_root / "src"; src.mkdir(parents=True, exist_ok=True)
    out = out_root / "out"; out.mkdir(parents=True, exist_ok=True)

    # subA — 色A/色C 数据全空(期待被压缩),色B/数量/单价 有值
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(['客户名', '型号', '类型', '色A', '色B', '色C', '数量', '单价', '辅助列', '换算率', '备注'])
    ws.append(['客户1', 'M001', '类型X', None, '红', None, 100, 50, 0, 1.0, 'note1'])
    ws.append(['客户2', 'M002', '类型X', None, '蓝', None, 200, 60, 1, 2.0, 'note2'])
    ws.append(['合计', '合计', '合计', None, None, None, 300, None, None, None, None])
    wb.save(src / 'subA.xlsx')

    # subB — 列顺序不同,色A(在 subA 是空)/色C(在 subA 也是空)在这张表里都空 → 整组同时空,可被压缩
    #         「备注/辅助列/辅助2」按表头名预剔除
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(['客户名', '型号', '类型', '色A', '色B', '色C', '数量', '备注', '辅助列', '辅助2'])
    ws.append(['客户3', 'M003', '类型Y', None, '绿', None, 50, 'nb', 0, 'x'])
    ws.append(['客户4', 'M004', '类型Y', None, '黄', None, 70, 'nb', 1, 'y'])
    ws.append(['合计', '合计', '合计', None, None, None, 120, None, None, None])
    wb.save(src / 'subB.xlsx')

    # 分类表
    idx = out_root / "idx.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(['组A', '描述', '子表1', '子表2'])
    ws.append(['组A', 'd', 'subA', 'subB'])
    wb.save(idx)

    return src, idx, out


def main(argv: list[str] | None = None) -> int:
    """CLI 入口,支持以下用法:

    默认(读 config.json):        python merge_stock_files.py
    显式传参:                    python merge_stock_files.py --src <dir> --index <file> --out <dir>
    一键生成演示数据并跑一次:    python merge_stock_files.py --demo
    演示数据生成但只测解析:      python merge_stock_files.py --demo --dry-run
    """
    import argparse

    parser = argparse.ArgumentParser(description="分类表 → 合并 Excel")
    parser.add_argument("--src", help="源子表目录(覆盖 config.json)")
    parser.add_argument("--index", help="分类表文件(覆盖 config.json)")
    parser.add_argument("--out", help="输出目录(覆盖 config.json)")
    parser.add_argument("--demo", action="store_true",
                        help="在 --demo-dir 指定的目录下生成演示数据并执行一次合并")
    parser.add_argument("--demo-dir", default=str(Path.home() / "stock_demo"),
                        help="演示数据生成目录(默认 ~/stock_demo)")
    parser.add_argument("--dry-run", action="store_true",
                        help="只生成演示数据 / 解析,不做实际合并(便于排错)")
    parser.add_argument("--unsafe-output", action="store_true",
                        help="允许 output 指向 src_dir 等危险路径(与 config 中相同选项等价)")
    args = parser.parse_args(argv)

    cfg: dict
    if args.demo:
        demo_dir = Path(args.demo_dir).resolve()
        src, idx, out = _demo_data(demo_dir)
        cfg = {
            "src_dir": str(src),
            "index_file": str(idx),
            "output_dir": str(out),
            "allow_unsafe_output": args.unsafe_output,
        }
        print(f"[demo] 数据已生成到 {demo_dir}")
        print(f"        src   = {src}")
        print(f"        index = {idx}")
        print(f"        out   = {out}")
    else:
        cfg = load_config()
        if args.src:
            cfg["src_dir"] = args.src
        if args.index:
            cfg["index_file"] = args.index
        if args.out:
            cfg["output_dir"] = args.out
        if args.unsafe_output:
            cfg["allow_unsafe_output"] = True

    print(f"SRC_DIR    = {cfg['src_dir']}")
    print(f"INDEX_FILE = {cfg['index_file']}")
    print(f"OUTPUT_DIR = {cfg['output_dir']}")
    print("NUM_COLS   = 自动(预剔除辅助列 + 移除全空数据列)")
    print()

    if args.dry_run:
        try:
            groups = load_classification(Path(cfg["index_file"]))
        except FileNotFoundError as e:
            print(f"[错误] {e}")
            return 1
        print(f"[dry-run] 分类表共 {len(groups)} 组")
        for g in groups:
            print(f"  组: {g['a_col']!r} -> 子表 {g['sub_files']}")
        return 0

    try:
        result = process(cfg, progress_cb=lambda c, t, m: None)
    except (ValueError, FileNotFoundError) as e:
        print(f"[错误] {e}")
        return 1

    print("\n".join(result.log))
    print()
    print(f"[完成] 合并 {result.groups_merged} / {result.groups_total} 组，"
          f"有效 {result.rows_total} 行，过滤 {result.rows_filtered_total} 行，"
          f"复制 {result.files_copied} 个源文件。")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
